
import streamlit as st
import pandas as pd
from timezonefinder import TimezoneFinder
from openpyxl import load_workbook
import pytz
from datetime import datetime
import io

st.set_page_config(page_title="Conversor IBUTG", layout="centered")

st.title("📊 Conversor de Planilhas IBUTG")
st.markdown("Este aplicativo converte arquivos do INMET para o modelo de planilha IBUTG automaticamente.")

modelo_file = st.file_uploader("📄 Envie a planilha modelo IBUTG (.xlsx)", type=["xlsx"])
csv_uploaded = st.file_uploader("📁 Envie os arquivos INMET (.csv, .CSV, .xls, .xlsx)", type=["csv", "CSV", "xlsx", "xls"], accept_multiple_files=True)

if modelo_file and csv_uploaded:
    modelo_bytes = modelo_file.read()
    for file in csv_uploaded:
        st.write(f"🔄 Processando: {file.name}")
        file_content = file.read().decode('latin1')
        linhas = file_content.splitlines()
        latitude = longitude = None
        for linha in linhas[:10]:
            if 'LATITUDE' in linha.upper():
                latitude = float(linha.split(':')[-1].replace(';', '').replace(',', '.'))
            if 'LONGITUDE' in linha.upper():
                longitude = float(linha.split(':')[-1].replace(';', '').replace(',', '.'))
        if latitude is None or longitude is None:
            st.error("❌ Latitude ou longitude não encontrada.")
            continue
        tf = TimezoneFinder()
        timezone = pytz.timezone(tf.timezone_at(lng=longitude, lat=latitude))
        df = pd.read_csv(io.StringIO(file_content), sep=';', skiprows=8, encoding='latin1')
        df = df[[
            'DATA (YYYY-MM-DD)', 'HORA (UTC)',
            'TEMPERATURA DO AR - BULBO SECO, HORARIA (°C)',
            'TEMPERATURA DO PONTO DE ORVALHO (°C)',
            'UMIDADE RELATIVA DO AR, HORARIA (%)',
            'VENTO, VELOCIDADE HORARIA (m/s)'
        ]]
        for col in df.columns[2:]:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '.'), errors='coerce')
        df['data_hora_utc'] = pd.to_datetime(df['DATA (YYYY-MM-DD)'] + ' ' + df['HORA (UTC)'], format='%Y-%m-%d %H:%M', utc=True)
        df['data_hora_local'] = df['data_hora_utc'].dt.tz_convert(timezone)
        df['hora_local'] = df['data_hora_local'].dt.hour
        df_filtrado = df[(df['hora_local'] >= 8) & (df['hora_local'] <= 17)].copy()
        df_filtrado = df_filtrado.sort_values(by='data_hora_local').reset_index(drop=True)
        dados_final = pd.DataFrame({
            'DATA': df_filtrado['data_hora_local'].dt.date,
            'HORA': df_filtrado['data_hora_local'].dt.strftime('%H:%M'),
            'TAR': df_filtrado['TEMPERATURA DO AR - BULBO SECO, HORARIA (°C)'],
            'TPO': df_filtrado['TEMPERATURA DO PONTO DE ORVALHO (°C)'],
            'UR': df_filtrado['UMIDADE RELATIVA DO AR, HORARIA (%)'],
            'VENTO': df_filtrado['VENTO, VELOCIDADE HORARIA (m/s)']
        })
        wb = load_workbook(io.BytesIO(modelo_bytes))
        ws = wb.active
        linha = 4
        for _, row in dados_final.iterrows():
            ws[f'D{linha}'] = row['DATA']
            ws[f'E{linha}'] = row['HORA']
            ws[f'G{linha}'] = row['TAR']
            ws[f'H{linha}'] = row['TPO']
            ws[f'I{linha}'] = row['UR']
            ws[f'J{linha}'] = row['VENTO']
            linha += 1
        output = io.BytesIO()
        wb.save(output)
        st.download_button("⬇️ Baixar planilha IBUTG convertida", data=output.getvalue(), file_name=f"IBUTG_{file.name.replace('.csv', '')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
